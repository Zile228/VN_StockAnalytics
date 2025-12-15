"""
Fundamentals ingestion utilities for `data/timeseries/bank.xlsx` (no pandas).

Goal:
- Normalize multi-sheet Excel into a canonical long-form table:
  (symbol, year, quarter, statement, metric, value)

Notes:
- `bank.xlsx` sheets may contain a "meta header row" above the actual header.
- Some sheets may have inflated max_row due to formatting; ingestion stops after
  consecutive empty rows threshold to avoid scanning millions of blank rows.
"""

from __future__ import annotations

import csv
import re
from dataclasses import dataclass
from pathlib import Path
from typing import Dict, Iterable, Iterator, List, Optional, Sequence, Tuple

from openpyxl import load_workbook


@dataclass(frozen=True, slots=True)
class FundamentalRecord:
    symbol: str
    year: int
    quarter: int
    statement: str
    metric: str
    metric_raw: str
    value: float


def _norm_metric(name: str) -> str:
    """
    Conservative normalization for metric names:
    - lower
    - replace non-alnum with underscore
    - collapse underscores
    Keeps unicode letters (Vietnamese) as-is.
    """
    s = (name or "").strip().lower()
    s = re.sub(r"[^\w]+", "_", s, flags=re.UNICODE)
    s = re.sub(r"_+", "_", s).strip("_")
    return s


def _to_int(v: object) -> Optional[int]:
    if v is None:
        return None
    try:
        s = str(v).strip()
        if s == "":
            return None
        # handle floats like 2025.0
        return int(float(s))
    except Exception:
        return None


def _to_float(v: object) -> Optional[float]:
    if v is None:
        return None
    try:
        s = str(v).strip()
        if s == "":
            return None
        return float(s.replace(",", ""))
    except Exception:
        return None


def _is_row_empty(row: Sequence[object]) -> bool:
    for x in row:
        if x is None:
            continue
        if str(x).strip() != "":
            return False
    return True


def _find_header_row(ws, *, search_rows: int = 50) -> Optional[Tuple[int, List[str]]]:
    """
    Find the row index containing 'CP' (ticker), 'Năm', 'Kỳ' header.
    Returns (row_idx_1based, header_list).
    """
    for i, row in enumerate(ws.iter_rows(min_row=1, max_row=search_rows, values_only=True), start=1):
        header = [str(x).strip() if x is not None else "" for x in row]
        # heuristic: must contain CP and Năm and Kỳ
        if "CP" in header and "Năm" in header and "Kỳ" in header:
            return (i, header)
    return None


def iter_bank_fundamentals_long(
    bank_xlsx_path: Path,
    *,
    sheets: Optional[Sequence[str]] = None,
    empty_row_break: int = 200,
) -> Iterator[FundamentalRecord]:
    """
    Yield normalized FundamentalRecord from bank.xlsx.
    """
    wb = load_workbook(bank_xlsx_path, read_only=True, data_only=True)
    use_sheets = list(sheets) if sheets else wb.sheetnames

    for sheet_name in use_sheets:
        ws = wb[sheet_name]
        found = _find_header_row(ws)
        if not found:
            continue
        header_row_idx, header = found
        col = {h: j for j, h in enumerate(header) if h}

        base_cols = {"CP", "Năm", "Kỳ"}
        metric_cols = [(j, h) for h, j in col.items() if h not in base_cols]

        empty_run = 0
        for row in ws.iter_rows(min_row=header_row_idx + 1, values_only=True):
            if _is_row_empty(row):
                empty_run += 1
                if empty_run >= empty_row_break:
                    break
                continue
            empty_run = 0

            sym = str(row[col["CP"]] if col["CP"] < len(row) else "").strip().upper()
            year = _to_int(row[col["Năm"]] if col["Năm"] < len(row) else None)
            quarter = _to_int(row[col["Kỳ"]] if col["Kỳ"] < len(row) else None)
            if not sym or year is None or quarter is None:
                continue

            for j, raw_name in metric_cols:
                if j >= len(row):
                    continue
                v = _to_float(row[j])
                if v is None:
                    continue
                yield FundamentalRecord(
                    symbol=sym,
                    year=year,
                    quarter=quarter,
                    statement=sheet_name,
                    metric=_norm_metric(raw_name),
                    metric_raw=raw_name,
                    value=v,
                )


def write_fundamentals_long_csv(
    bank_xlsx_path: Path,
    out_csv_path: Path,
    *,
    sheets: Optional[Sequence[str]] = None,
    empty_row_break: int = 200,
) -> Dict[str, int]:
    """
    Convert bank.xlsx into canonical long-form CSV.
    Returns simple stats.
    """
    out_csv_path.parent.mkdir(parents=True, exist_ok=True)
    fieldnames = ["symbol", "year", "quarter", "statement", "metric", "metric_raw", "value"]
    n = 0
    by_sheet: Dict[str, int] = {}

    with out_csv_path.open("w", encoding="utf-8", newline="") as f:
        w = csv.DictWriter(f, fieldnames=fieldnames)
        w.writeheader()
        for rec in iter_bank_fundamentals_long(
            bank_xlsx_path, sheets=sheets, empty_row_break=empty_row_break
        ):
            w.writerow(
                {
                    "symbol": rec.symbol,
                    "year": rec.year,
                    "quarter": rec.quarter,
                    "statement": rec.statement,
                    "metric": rec.metric,
                    "metric_raw": rec.metric_raw,
                    "value": rec.value,
                }
            )
            n += 1
            by_sheet[rec.statement] = by_sheet.get(rec.statement, 0) + 1

    return {"rows": n, **{f"rows_{k}": v for k, v in sorted(by_sheet.items())}}


