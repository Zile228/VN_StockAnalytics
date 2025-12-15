"""
Data Quality Report generator (no pandas).

Scans local datasets in data/ and produces:
- machine-readable JSON
- human-readable Markdown

Focus: counts, ranges, missing, anomalies, schema/parse issues.
"""

from __future__ import annotations

import csv
import dataclasses
import datetime as dt
import json
import math
from collections import Counter, defaultdict
from dataclasses import dataclass
from pathlib import Path
from typing import Any, Dict, Iterable, List, Optional, Tuple

from openpyxl import load_workbook


def _now_ts() -> str:
    return dt.datetime.now().strftime("%Y%m%d_%H%M%S")


def _try_parse_iso_dt(s: str) -> Optional[dt.datetime]:
    raw = (s or "").strip()
    if not raw:
        return None
    raw2 = raw.replace("T", " ")
    if len(raw2) == 16:
        raw2 = raw2 + ":00"
    for fmt in ("%Y-%m-%d %H:%M:%S", "%Y-%m-%d %H:%M"):
        try:
            return dt.datetime.strptime(raw2, fmt)
        except ValueError:
            pass
    try:
        return dt.datetime.fromisoformat(raw)
    except Exception:
        return None


def _try_parse_date_ymd(s: str) -> Optional[dt.date]:
    raw = (s or "").strip()
    if not raw:
        return None
    try:
        return dt.date.fromisoformat(raw[:10])
    except Exception:
        return None


def _try_parse_date_dmy(s: str) -> Optional[dt.date]:
    raw = (s or "").strip()
    if not raw:
        return None
    # DD/MM/YYYY
    try:
        return dt.datetime.strptime(raw, "%d/%m/%Y").date()
    except Exception:
        return None


def _parse_float_vn(s: str) -> Optional[float]:
    raw = (s or "").strip()
    if raw == "":
        return None
    raw = raw.replace(",", "")
    try:
        return float(raw)
    except Exception:
        return None


def _parse_percent_vn(s: str) -> Optional[float]:
    raw = (s or "").strip()
    if raw == "":
        return None
    raw = raw.replace("%", "").replace(",", ".")
    try:
        return float(raw) / 100.0
    except Exception:
        return None


def _minmax_update(v: Any, cur_min: Any, cur_max: Any) -> tuple[Any, Any]:
    if v is None:
        return cur_min, cur_max
    if cur_min is None or v < cur_min:
        cur_min = v
    if cur_max is None or v > cur_max:
        cur_max = v
    return cur_min, cur_max


def _is_row_empty(row: Iterable[object]) -> bool:
    for x in row:
        if x is None:
            continue
        if str(x).strip() != "":
            return False
    return True


def scan_vn30_history(path: Path) -> Dict[str, Any]:
    """
    vn30_history.csv: time,open,high,low,close,volume,symbol
    """
    by_sym = defaultdict(lambda: {"n": 0, "min_date": None, "max_date": None, "dups": 0})
    seen = defaultdict(set)
    anomalies = Counter()
    bad_rows = 0
    with path.open("r", encoding="utf-8-sig", newline="") as f:
        r = csv.DictReader(f)
        for row in r:
            sym = (row.get("symbol") or "").strip().upper()
            d = (row.get("time") or "").strip()
            if not sym or not d:
                bad_rows += 1
                anomalies["missing_symbol_or_time"] += 1
                continue

            dd = _try_parse_date_ymd(d)
            if dd is None:
                bad_rows += 1
                anomalies["bad_date_format"] += 1
                continue

            key = dd.isoformat()
            by_sym[sym]["n"] += 1
            if key in seen[sym]:
                by_sym[sym]["dups"] += 1
            else:
                seen[sym].add(key)

            mn, mx = by_sym[sym]["min_date"], by_sym[sym]["max_date"]
            mn, mx = _minmax_update(dd, mn, mx)
            by_sym[sym]["min_date"], by_sym[sym]["max_date"] = mn, mx

            o = _parse_float_vn(row.get("open", ""))
            h = _parse_float_vn(row.get("high", ""))
            l = _parse_float_vn(row.get("low", ""))
            c = _parse_float_vn(row.get("close", ""))
            v = _parse_float_vn(row.get("volume", ""))
            if any(x is None for x in (o, h, l, c, v)):
                anomalies["missing_numeric_fields"] += 1
                continue
            if v is not None and v < 0:
                anomalies["negative_volume"] += 1
            if c is not None and c <= 0:
                anomalies["non_positive_close"] += 1
            if h is not None and l is not None and h < l:
                anomalies["high_lt_low"] += 1
            if h is not None and o is not None and c is not None and h < max(o, c):
                anomalies["high_lt_max_open_close"] += 1
            if l is not None and o is not None and c is not None and l > min(o, c):
                anomalies["low_gt_min_open_close"] += 1

    syms = sorted(by_sym.keys())
    return {
        "file": str(path),
        "rows_by_symbol": {s: {**v, "min_date": (v["min_date"].isoformat() if v["min_date"] else None), "max_date": (v["max_date"].isoformat() if v["max_date"] else None)} for s, v in by_sym.items()},
        "symbol_count": len(syms),
        "total_rows": int(sum(by_sym[s]["n"] for s in syms)),
        "bad_rows": int(bad_rows),
        "anomalies": dict(anomalies),
    }


def scan_news_bronze(path: Path) -> Dict[str, Any]:
    counts = Counter()
    sym_counts = Counter()
    min_dt: Optional[dt.datetime] = None
    max_dt: Optional[dt.datetime] = None
    bad_dt = 0
    multi_symbol_rows = 0
    missing_fields = Counter()

    with path.open("r", encoding="utf-8-sig", newline="") as f:
        r = csv.DictReader(f)
        for row in r:
            counts["rows"] += 1
            sym = (row.get("symbol") or "").strip()
            if "," in sym:
                multi_symbol_rows += 1
            if not sym:
                missing_fields["symbol"] += 1
            else:
                sym_counts[sym.upper()] += 1

            for fld in ("date_time", "title", "source"):
                if not (row.get(fld) or "").strip():
                    missing_fields[fld] += 1

            t = _try_parse_iso_dt(row.get("date_time", "") or "")
            if t is None:
                bad_dt += 1
            else:
                min_dt, max_dt = _minmax_update(t, min_dt, max_dt)

    return {
        "file": str(path),
        "total_rows": int(counts["rows"]),
        "unique_symbols": int(len(sym_counts)),
        "time_range": {"min": min_dt.isoformat() if min_dt else None, "max": max_dt.isoformat() if max_dt else None},
        "bad_datetime_rows": int(bad_dt),
        "multi_symbol_rows": int(multi_symbol_rows),
        "missing_fields": dict(missing_fields),
        "top_symbols": sym_counts.most_common(10),
    }


def scan_sentiment_silver(path: Path) -> Dict[str, Any]:
    counts = Counter()
    sym_counts = Counter()
    min_dt: Optional[dt.datetime] = None
    max_dt: Optional[dt.datetime] = None
    bad_dt = 0
    missing_fields = Counter()
    bad_numeric = Counter()
    model_counts = Counter()

    with path.open("r", encoding="utf-8-sig", newline="") as f:
        r = csv.DictReader(f)
        for row in r:
            counts["rows"] += 1
            sym = (row.get("symbol") or "").strip().upper()
            if not sym:
                missing_fields["symbol"] += 1
            else:
                sym_counts[sym] += 1

            # datetime: date_dt preferred
            t = _try_parse_iso_dt((row.get("date_dt") or row.get("date_time") or ""))
            if t is None:
                bad_dt += 1
            else:
                min_dt, max_dt = _minmax_update(t, min_dt, max_dt)

            # required analytics
            sc = (row.get("sentiment_score") or "").strip()
            rel = (row.get("relevance") or "").strip()
            if sc == "":
                missing_fields["sentiment_score"] += 1
            else:
                try:
                    _ = int(float(sc))
                except Exception:
                    bad_numeric["sentiment_score"] += 1
            if rel == "":
                missing_fields["relevance"] += 1
            else:
                try:
                    x = float(rel)
                    if x < 0 or x > 1:
                        bad_numeric["relevance_out_of_range"] += 1
                except Exception:
                    bad_numeric["relevance"] += 1

            mu = (row.get("model_used") or "").strip()
            if mu:
                model_counts[mu] += 1

    return {
        "file": str(path),
        "total_rows": int(counts["rows"]),
        "unique_symbols": int(len(sym_counts)),
        "time_range": {"min": min_dt.isoformat() if min_dt else None, "max": max_dt.isoformat() if max_dt else None},
        "bad_datetime_rows": int(bad_dt),
        "missing_fields": dict(missing_fields),
        "bad_numeric": dict(bad_numeric),
        "top_symbols": sym_counts.most_common(10),
        "model_used_counts": model_counts.most_common(10),
    }


def scan_macro_xlsx(path: Path) -> Dict[str, Any]:
    if not path.exists():
        return {"file": str(path), "exists": False}
    wb = load_workbook(path, read_only=True, data_only=True)
    sheets = wb.sheetnames
    ws = wb[sheets[0]]
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return {"file": str(path), "exists": True, "sheets": sheets, "rows": 0}
    header = [str(x).strip() if x is not None else "" for x in rows[0]]
    col = {name: i for i, name in enumerate(header)}
    needed = {"Year", "Quarter", "INF", "GDP", "DC"}
    missing = sorted(list(needed - set(col.keys())))
    parsed = 0
    bad = 0
    inf_rng = (None, None)
    gdp_rng = (None, None)
    dc_rng = (None, None)
    for r in rows[1:]:
        try:
            year = int(str(r[col["Year"]]).strip())
            q = int(str(r[col["Quarter"]]).strip())
        except Exception:
            bad += 1
            continue
        inf = _parse_percent_vn(str(r[col["INF"]]) if "INF" in col else "")
        gdp = _parse_percent_vn(str(r[col["GDP"]]) if "GDP" in col else "")
        dc = _parse_percent_vn(str(r[col["DC"]]) if "DC" in col else "")
        parsed += 1
        inf_rng = _minmax_update(inf, inf_rng[0], inf_rng[1])
        gdp_rng = _minmax_update(gdp, gdp_rng[0], gdp_rng[1])
        dc_rng = _minmax_update(dc, dc_rng[0], dc_rng[1])

    return {
        "file": str(path),
        "exists": True,
        "sheets": sheets,
        "rows": len(rows) - 1,
        "missing_expected_columns": missing,
        "parsed_rows": int(parsed),
        "bad_rows": int(bad),
        "ranges": {
            "INF": {"min": inf_rng[0], "max": inf_rng[1]},
            "GDP": {"min": gdp_rng[0], "max": gdp_rng[1]},
            "DC": {"min": dc_rng[0], "max": dc_rng[1]},
        },
    }


def scan_usdvnd_csv(path: Path) -> Dict[str, Any]:
    if not path.exists():
        return {"file": str(path), "exists": False}
    with path.open("r", encoding="utf-8-sig", newline="") as f:
        r = csv.DictReader(f)
        rows = 0
        bad_date = 0
        bad_close = 0
        min_d: Optional[dt.date] = None
        max_d: Optional[dt.date] = None
        close_rng = (None, None)
        pct_rng = (None, None)
        for row in r:
            rows += 1
            d = _try_parse_date_dmy((row.get("Ngày") or "").strip())
            if d is None:
                bad_date += 1
                continue
            min_d, max_d = _minmax_update(d, min_d, max_d)
            close = _parse_float_vn((row.get("Lần cuối") or "").strip())
            if close is None:
                bad_close += 1
            else:
                close_rng = _minmax_update(close, close_rng[0], close_rng[1])
            pct = _parse_percent_vn((row.get("% Thay đổi") or "").strip())
            pct_rng = _minmax_update(pct, pct_rng[0], pct_rng[1])

    return {
        "file": str(path),
        "exists": True,
        "rows": int(rows),
        "date_range": {"min": min_d.isoformat() if min_d else None, "max": max_d.isoformat() if max_d else None},
        "bad_date_rows": int(bad_date),
        "bad_close_rows": int(bad_close),
        "ranges": {
            "close": {"min": close_rng[0], "max": close_rng[1]},
            "pct_change": {"min": pct_rng[0], "max": pct_rng[1]},
        },
    }


def scan_bank_xlsx(path: Path) -> Dict[str, Any]:
    if not path.exists():
        return {"file": str(path), "exists": False}
    wb = load_workbook(path, read_only=True, data_only=True)
    out: Dict[str, Any] = {"file": str(path), "exists": True, "sheets": wb.sheetnames, "per_sheet": {}}

    for name in wb.sheetnames:
        ws = wb[name]
        # Find actual header row (some sheets have meta header above)
        header_row_idx = None
        header_s: List[str] = []
        for ridx, row in enumerate(ws.iter_rows(min_row=1, max_row=50, values_only=True), start=1):
            row_s = [str(x).strip() if x is not None else "" for x in row]
            if "CP" in row_s and "Năm" in row_s and "Kỳ" in row_s:
                header_row_idx = ridx
                header_s = row_s
                break

        if header_row_idx is None or not header_s:
            out["per_sheet"][name] = {"rows_scanned": 0, "header_row": None, "bad": {"header_not_found": 1}}
            continue

        col = {h: i for i, h in enumerate(header_s) if h}
        rows_scanned = 0
        empty_run = 0
        sym_counts = Counter()
        bad = Counter()

        for r in ws.iter_rows(min_row=header_row_idx + 1, values_only=True):
            # stop scanning when hitting many consecutive empty rows (avoid huge max_row from formatting)
            if _is_row_empty(r):
                empty_run += 1
                if empty_run >= 200:
                    break
                continue
            empty_run = 0
            rows_scanned += 1

            if "CP" in col:
                sym = str(r[col["CP"]] if col["CP"] < len(r) else "" or "").strip().upper()
                if sym:
                    sym_counts[sym] += 1
                else:
                    bad["missing_symbol"] += 1
            for fld in ("Năm", "Kỳ"):
                if fld in col:
                    v = r[col[fld]] if col[fld] < len(r) else None
                    if v is None or str(v).strip() == "":
                        bad[f"missing_{fld}"] += 1

        out["per_sheet"][name] = {
            "header_row": header_row_idx,
            "rows_scanned": int(rows_scanned),
            "columns": header_s[:60],
            "unique_symbols": int(len(sym_counts)) if sym_counts else 0,
            "top_symbols": sym_counts.most_common(10),
            "bad": dict(bad),
        }
    return out


def build_report(*, paths: Dict[str, Path]) -> Dict[str, Any]:
    """
    paths keys:
      - vn30_history_csv
      - news_csv
      - sentiment_csv
      - macro_xlsx
      - usdvnd_csv
      - bank_xlsx (optional)
    """
    report: Dict[str, Any] = {
        "generated_at": dt.datetime.now().isoformat(),
        "datasets": {},
    }
    report["datasets"]["vn30_history"] = scan_vn30_history(paths["vn30_history_csv"])
    report["datasets"]["news_bronze"] = scan_news_bronze(paths["news_csv"])
    report["datasets"]["sentiment_silver"] = scan_sentiment_silver(paths["sentiment_csv"])
    report["datasets"]["macro"] = scan_macro_xlsx(paths["macro_xlsx"])
    report["datasets"]["usdvnd"] = scan_usdvnd_csv(paths["usdvnd_csv"])
    if "bank_xlsx" in paths:
        report["datasets"]["bank"] = scan_bank_xlsx(paths["bank_xlsx"])
    return report


def render_markdown(report: Dict[str, Any]) -> str:
    ds = report.get("datasets", {})
    lines: List[str] = []
    lines.append("## Data Quality Report")
    lines.append("")
    lines.append(f"- **generated_at**: `{report.get('generated_at')}`")
    lines.append("")

    def section(title: str) -> None:
        lines.append(f"### {title}")
        lines.append("")

    # vn30 history
    h = ds.get("vn30_history", {})
    section("Market OHLCV (vn30_history.csv)")
    lines.append(f"- **file**: `{h.get('file')}`")
    lines.append(f"- **symbols**: {h.get('symbol_count')}")
    lines.append(f"- **total_rows**: {h.get('total_rows')}")
    lines.append(f"- **bad_rows**: {h.get('bad_rows')}")
    lines.append(f"- **anomalies**: `{json.dumps(h.get('anomalies', {}), ensure_ascii=False)}`")
    lines.append("")

    # news
    n = ds.get("news_bronze", {})
    section("News Bronze (VN30_Merged_News.csv)")
    lines.append(f"- **file**: `{n.get('file')}`")
    lines.append(f"- **total_rows**: {n.get('total_rows')}")
    lines.append(f"- **unique_symbols**: {n.get('unique_symbols')}")
    tr = n.get("time_range", {}) or {}
    lines.append(f"- **time_range**: `{tr.get('min')}` → `{tr.get('max')}`")
    lines.append(f"- **bad_datetime_rows**: {n.get('bad_datetime_rows')}")
    lines.append(f"- **multi_symbol_rows**: {n.get('multi_symbol_rows')}")
    lines.append(f"- **missing_fields**: `{json.dumps(n.get('missing_fields', {}), ensure_ascii=False)}`")
    lines.append("")

    # sentiment
    s = ds.get("sentiment_silver", {})
    section("Sentiment Silver (VN30_Sentiment_Analyzed.csv)")
    lines.append(f"- **file**: `{s.get('file')}`")
    lines.append(f"- **total_rows**: {s.get('total_rows')}")
    lines.append(f"- **unique_symbols**: {s.get('unique_symbols')}")
    tr = s.get("time_range", {}) or {}
    lines.append(f"- **time_range**: `{tr.get('min')}` → `{tr.get('max')}`")
    lines.append(f"- **bad_datetime_rows**: {s.get('bad_datetime_rows')}")
    lines.append(f"- **missing_fields**: `{json.dumps(s.get('missing_fields', {}), ensure_ascii=False)}`")
    lines.append(f"- **bad_numeric**: `{json.dumps(s.get('bad_numeric', {}), ensure_ascii=False)}`")
    lines.append("")

    # macro
    m = ds.get("macro", {})
    section("Macro (macro.xlsx)")
    lines.append(f"- **file**: `{m.get('file')}`")
    lines.append(f"- **exists**: {m.get('exists')}")
    if m.get("exists"):
        lines.append(f"- **sheets**: `{m.get('sheets')}`")
        lines.append(f"- **rows**: {m.get('rows')}")
        lines.append(f"- **missing_expected_columns**: `{m.get('missing_expected_columns')}`")
        lines.append(f"- **ranges**: `{json.dumps(m.get('ranges', {}), ensure_ascii=False)}`")
    lines.append("")

    # usdvnd
    u = ds.get("usdvnd", {})
    section("FX (USDVND)")
    lines.append(f"- **file**: `{u.get('file')}`")
    lines.append(f"- **exists**: {u.get('exists')}")
    if u.get("exists"):
        dr = u.get("date_range", {}) or {}
        lines.append(f"- **rows**: {u.get('rows')}")
        lines.append(f"- **date_range**: `{dr.get('min')}` → `{dr.get('max')}`")
        lines.append(f"- **bad_date_rows**: {u.get('bad_date_rows')}")
        lines.append(f"- **bad_close_rows**: {u.get('bad_close_rows')}")
        lines.append(f"- **ranges**: `{json.dumps(u.get('ranges', {}), ensure_ascii=False)}`")
    lines.append("")

    # bank
    b = ds.get("bank")
    if b:
        section("Fundamental (bank.xlsx)")
        lines.append(f"- **file**: `{b.get('file')}`")
        lines.append(f"- **exists**: {b.get('exists')}")
        if b.get("exists"):
            lines.append(f"- **sheets**: `{b.get('sheets')}`")
            per = b.get("per_sheet", {}) or {}
            for sh, meta in per.items():
                lines.append(f"  - **{sh}**: rows={meta.get('rows')}, unique_symbols={meta.get('unique_symbols')}, bad={meta.get('bad')}")
        lines.append("")

    return "\n".join(lines) + "\n"


