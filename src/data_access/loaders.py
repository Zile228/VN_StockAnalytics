"""
Local data loaders for VN_StockAnalytics.

Constraints:
- Prefer stdlib csv/datetime + numpy over pandas.
- Use openpyxl for xlsx.
"""

from __future__ import annotations

import csv
import datetime as dt
from dataclasses import dataclass
from pathlib import Path
from typing import Dict, Iterable, List, Optional, Tuple

from openpyxl import load_workbook


@dataclass(frozen=True, slots=True)
class OHLCVBar:
    date: dt.date
    open: float
    high: float
    low: float
    close: float
    volume: float


@dataclass(frozen=True, slots=True)
class NewsItem:
    symbol: str
    published_at: dt.datetime
    title: str
    sapo: str
    url: str
    source: str


@dataclass(frozen=True, slots=True)
class SentimentItem:
    symbol: str
    published_at: dt.datetime
    sentiment_score: float
    relevance: float
    title: str
    reasoning: str
    source: str
    url: str


@dataclass(frozen=True, slots=True)
class MacroPoint:
    year: int
    quarter: int
    inf_pct: Optional[float]
    gdp_pct: Optional[float]
    dc_pct: Optional[float]


def _parse_date(s: str) -> dt.date:
    return dt.date.fromisoformat(s.strip())


def _parse_datetime_flexible(s: str) -> dt.datetime:
    raw = s.strip()
    # Examples observed:
    # - "2025-12-08 16:41"
    # - "2025-12-10T08:01:00"
    # - "2025-12-08 16:41:00"
    fmts = (
        "%Y-%m-%d %H:%M",
        "%Y-%m-%d %H:%M:%S",
        "%Y-%m-%dT%H:%M:%S",
        "%Y-%m-%dT%H:%M",
    )
    for f in fmts:
        try:
            return dt.datetime.strptime(raw, f)
        except ValueError:
            continue
    # last resort: try fromisoformat (handles many variants)
    try:
        return dt.datetime.fromisoformat(raw)
    except ValueError as e:
        raise ValueError(f"Unrecognized datetime format: {s!r}") from e


def _parse_float(s: str) -> float:
    raw = (s or "").strip()
    if raw == "":
        return float("nan")
    return float(raw.replace(",", ""))


def _parse_percent_str(s: object) -> Optional[float]:
    """
    Parse strings like '2,95%' into 0.0295.
    Returns None for blanks/None.
    """
    if s is None:
        return None
    raw = str(s).strip()
    if raw == "":
        return None
    raw = raw.replace("%", "").replace(",", ".")
    try:
        return float(raw) / 100.0
    except ValueError:
        return None


def load_vn30_history_csv(path: Path) -> Dict[str, List[OHLCVBar]]:
    """
    Expected header: time,open,high,low,close,volume,symbol
    """
    out: Dict[str, List[OHLCVBar]] = {}
    with path.open("r", encoding="utf-8-sig", newline="") as f:
        reader = csv.DictReader(f)
        required = {"time", "open", "high", "low", "close", "volume", "symbol"}
        if not reader.fieldnames or not required.issubset(set(reader.fieldnames)):
            raise ValueError(f"vn30_history missing required columns: {required}")

        for row in reader:
            sym = (row.get("symbol") or "").strip().upper()
            if not sym:
                continue
            bar = OHLCVBar(
                date=_parse_date(row["time"]),
                open=_parse_float(row["open"]),
                high=_parse_float(row["high"]),
                low=_parse_float(row["low"]),
                close=_parse_float(row["close"]),
                volume=_parse_float(row["volume"]),
            )
            out.setdefault(sym, []).append(bar)

    # Ensure sort by date ascending per symbol
    for sym in list(out.keys()):
        out[sym].sort(key=lambda b: b.date)
        if not out[sym]:
            del out[sym]
    return out


def load_news_csv(path: Path) -> List[NewsItem]:
    """
    Expected header: symbol,date_time,title,sapo,content_url,source
    """
    items: List[NewsItem] = []
    with path.open("r", encoding="utf-8-sig", newline="") as f:
        reader = csv.DictReader(f)
        required = {"symbol", "date_time", "title", "sapo", "content_url", "source"}
        if not reader.fieldnames or not required.issubset(set(reader.fieldnames)):
            raise ValueError(f"news file missing required columns: {required}")
        for row in reader:
            sym = (row.get("symbol") or "").strip().upper()
            if not sym:
                continue
            items.append(
                NewsItem(
                    symbol=sym,
                    published_at=_parse_datetime_flexible(row["date_time"]),
                    title=(row.get("title") or "").strip(),
                    sapo=(row.get("sapo") or "").strip(),
                    url=(row.get("content_url") or "").strip(),
                    source=(row.get("source") or "").strip(),
                )
            )
    items.sort(key=lambda x: x.published_at)
    return items


def load_sentiment_analyzed_csv(path: Path) -> List[SentimentItem]:
    """
    Expected columns (subset ok):
    symbol, date_time, sentiment_score, relevance, title, reasoning, source, content_url
    """
    items: List[SentimentItem] = []
    with path.open("r", encoding="utf-8-sig", newline="") as f:
        reader = csv.DictReader(f)
        if not reader.fieldnames:
            return items
        for row in reader:
            sym = (row.get("symbol") or "").strip().upper()
            if not sym:
                continue
            date_time = row.get("date_dt") or row.get("date_time") or ""
            items.append(
                SentimentItem(
                    symbol=sym,
                    published_at=_parse_datetime_flexible(date_time),
                    sentiment_score=_parse_float(str(row.get("sentiment_score") or "")),
                    relevance=_parse_float(str(row.get("relevance") or "")),
                    title=(row.get("title") or "").strip(),
                    reasoning=(row.get("reasoning") or "").strip(),
                    source=(row.get("source") or "").strip(),
                    url=(row.get("content_url") or "").strip(),
                )
            )
    items.sort(key=lambda x: x.published_at)
    return items


def load_macro_xlsx(path: Path) -> List[MacroPoint]:
    """
    Reads Sheet1 with header: Year, Quarter, INF, GDP, DC (as % strings).
    Returns rates as decimals, e.g. '2,95%' -> 0.0295.
    """
    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb[wb.sheetnames[0]]
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return []
    header = [str(c).strip() if c is not None else "" for c in rows[0]]
    col = {name: idx for idx, name in enumerate(header)}
    needed = {"Year", "Quarter", "INF", "GDP", "DC"}
    if not needed.issubset(set(col.keys())):
        # tolerate different capitalization
        # fallback: return empty
        return []

    out: List[MacroPoint] = []
    for r in rows[1:]:
        try:
            year = int(str(r[col["Year"]]).strip())
            quarter = int(str(r[col["Quarter"]]).strip())
        except Exception:
            continue
        out.append(
            MacroPoint(
                year=year,
                quarter=quarter,
                inf_pct=_parse_percent_str(r[col["INF"]]),
                gdp_pct=_parse_percent_str(r[col["GDP"]]),
                dc_pct=_parse_percent_str(r[col["DC"]]),
            )
        )
    return out


def load_usdvnd_csv(path: Path) -> List[Tuple[dt.date, float]]:
    """
    Load USD/VND historical CSV. File name suggests Vietnamese encoding; columns may vary.
    Heuristic: find first date-like column and first numeric column.
    Returns list of (date, value) sorted ascending.
    """
    with path.open("r", encoding="utf-8-sig", newline="") as f:
        reader = csv.reader(f)
        header = next(reader, None)
        if header is None:
            return []
        # read few rows to infer indices
        sample: List[List[str]] = []
        for _ in range(20):
            row = next(reader, None)
            if row is None:
                break
            sample.append(row)

    # guess indices
    def looks_like_date(x: str) -> bool:
        try:
            _ = dt.date.fromisoformat(x.strip()[:10])
            return True
        except Exception:
            return False

    date_idx = None
    val_idx = None
    for j in range(len(header)):
        if any(looks_like_date(r[j]) for r in sample if j < len(r)):
            date_idx = j
            break
    for j in range(len(header)):
        if j == date_idx:
            continue
        try:
            _ = _parse_float(sample[0][j])
            val_idx = j
            break
        except Exception:
            continue

    if date_idx is None or val_idx is None:
        return []

    out: List[Tuple[dt.date, float]] = []
    with path.open("r", encoding="utf-8-sig", newline="") as f:
        reader = csv.reader(f)
        _ = next(reader, None)
        for row in reader:
            if date_idx >= len(row) or val_idx >= len(row):
                continue
            ds = row[date_idx].strip()
            vs = row[val_idx].strip()
            if not ds or not vs:
                continue
            try:
                d = dt.date.fromisoformat(ds[:10])
                v = _parse_float(vs)
            except Exception:
                continue
            out.append((d, v))
    out.sort(key=lambda x: x[0])
    return out


